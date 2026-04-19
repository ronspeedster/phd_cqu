# section 1

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import csv
import logging
import shutil
import subprocess
from collections import Counter
from typing import Iterable, List, Tuple

try:
    import matplotlib.pyplot as plt
except Exception:  # pragma: no cover
    plt = None

try:
    from dna_features_viewer import GraphicFeature, GraphicRecord
except Exception:  # pragma: no cover
    GraphicFeature = None
    GraphicRecord = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except Exception:  # pragma: no cover
    Workbook = None
    PatternFill = None


# section 2

IUPAC_ALLOWED = set("ACGTRYSWKMBDHVN-")
DNA_STRICT_ALLOWED = set("ACGT-")
LOGGER = logging.getLogger("hiv_preprocessing")


@dataclass
class SequenceRecord:
    header: str
    sequence: str


@dataclass
class PipelineConfig:
    input_fasta: Path
    output_dir: Path
    max_sequences_for_test: int = 0
    force_mafft_alignment: bool = False
    mafft_binary: str = "mafft"
    mafft_threads: int = -1
    consensus_name: str = "Subtype_A_Consensus"
    consensus_tie_char: str = "N"
    preview_window_start: int | None = None
    preview_window_size: int = 120
    preview_num_sequences: int = 12
    write_graphical_previews: bool = True
    write_excel_output: bool = True
    excel_filename: str = "consensus_and_alignment.xlsx"
    log_level: str = "INFO"


# section 3

def read_fasta(file_path: Path) -> List[SequenceRecord]:
    records: List[SequenceRecord] = []
    header = None
    seq_parts: List[str] = []

    with file_path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    records.append(SequenceRecord(header=header, sequence="".join(seq_parts)))
                header = line[1:].strip()
                seq_parts = []
            else:
                seq_parts.append(line)

    if header is not None:
        records.append(SequenceRecord(header=header, sequence="".join(seq_parts)))

    return records


def write_fasta(records: Iterable[SequenceRecord], file_path: Path, line_width: int = 80) -> None:
    with file_path.open("w", encoding="utf-8") as handle:
        for rec in records:
            handle.write(f">{rec.header}\n")
            seq = rec.sequence
            for i in range(0, len(seq), line_width):
                handle.write(seq[i : i + line_width] + "\n")


def normalize_sequence(seq: str) -> str:
    seq = seq.upper().replace("U", "T")
    # Keep only A/C/G/T/- for downstream Hypermut3 partial compatibility.
    normalized = [base if base in DNA_STRICT_ALLOWED else "-" for base in seq]
    return "".join(normalized)


def ungapped_length(seq: str) -> int:
    return sum(1 for c in seq if c != "-")


def n_fraction(seq: str) -> float:
    ungapped = [c for c in seq if c != "-"]
    if not ungapped:
        return 1.0
    n_count = sum(1 for c in ungapped if c == "N")
    return n_count / len(ungapped)


def is_aligned(records: List[SequenceRecord]) -> bool:
    if not records:
        return False
    lengths = {len(r.sequence) for r in records}
    return len(lengths) == 1


# section 4

def clean_records(records: List[SequenceRecord], cfg: PipelineConfig) -> Tuple[List[SequenceRecord], List[dict]]:
    cleaned: List[SequenceRecord] = []
    qc_rows: List[dict] = []

    for rec in records:
        norm_seq = normalize_sequence(rec.sequence)
        ungapped = ungapped_length(norm_seq)
        frac_n = n_fraction(norm_seq)

        qc_rows.append(
            {
                "header": rec.header,
                "raw_length": len(rec.sequence),
                "ungapped_length": ungapped,
                "n_fraction": round(frac_n, 6),
                "status": "kept_no_filtering",
            }
        )

        cleaned.append(SequenceRecord(header=rec.header, sequence=norm_seq))

    return cleaned, qc_rows


def write_qc_csv(rows: List[dict], file_path: Path) -> None:
    if not rows:
        return
    with file_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def align_with_mafft(records: List[SequenceRecord], cfg: PipelineConfig, tmp_input: Path, out_path: Path) -> List[SequenceRecord]:
    LOGGER.info("Step: writing temporary MAFFT input FASTA to %s", tmp_input)
    write_fasta(records, tmp_input)

    mafft_exe = shutil.which(cfg.mafft_binary)
    if not mafft_exe:
        raise RuntimeError(
            "Alignment required but MAFFT was not found on PATH. "
            "Install MAFFT or provide already aligned input."
        )

    cmd = [
        mafft_exe,
        "--thread",
        str(cfg.mafft_threads),
        "--auto",
        str(tmp_input),
    ]

    LOGGER.info("Step: running MAFFT alignment on %d sequences", len(records))

    proc = subprocess.run(cmd, text=True, capture_output=True, check=False)
    if proc.returncode != 0:
        raise RuntimeError(f"MAFFT failed with code {proc.returncode}: {proc.stderr}")

    out_path.write_text(proc.stdout, encoding="utf-8")
    aligned_records = read_fasta(out_path)
    LOGGER.info("Step complete: MAFFT alignment written to %s", out_path)
    return [
        SequenceRecord(header=rec.header, sequence=normalize_sequence(rec.sequence))
        for rec in aligned_records
    ]


# section 5

def iupac_for_bases(base_set: set[str]) -> str:
    mapping = {
        frozenset({"A"}): "A",
        frozenset({"C"}): "C",
        frozenset({"G"}): "G",
        frozenset({"T"}): "T",
        frozenset({"A", "G"}): "R",
        frozenset({"C", "T"}): "Y",
        frozenset({"G", "C"}): "S",
        frozenset({"A", "T"}): "W",
        frozenset({"G", "T"}): "K",
        frozenset({"A", "C"}): "M",
        frozenset({"C", "G", "T"}): "B",
        frozenset({"A", "G", "T"}): "D",
        frozenset({"A", "C", "T"}): "H",
        frozenset({"A", "C", "G"}): "V",
        frozenset({"A", "C", "G", "T"}): "N",
    }
    return mapping.get(frozenset(base_set), "N")


def consensus_from_alignment(records: List[SequenceRecord], tie_char: str = "N") -> str:
    if not records:
        raise ValueError("Cannot make consensus from empty record list")
    if not is_aligned(records):
        raise ValueError("Consensus requires aligned records of equal length")

    length = len(records[0].sequence)
    consensus_chars: List[str] = []

    for i in range(length):
        col = [r.sequence[i] for r in records]
        bases = [b for b in col if b in {"A", "C", "G", "T"}]
        if not bases:
            consensus_chars.append("-")
            continue

        counts = Counter(bases)
        top_count = max(counts.values())
        top_bases = {base for base, count in counts.items() if count == top_count}

        if len(top_bases) == 1:
            consensus_chars.append(next(iter(top_bases)))
        else:
            # Resolve ties without introducing ambiguous IUPAC letters.
            for candidate in ["A", "C", "G", "T"]:
                if candidate in top_bases:
                    consensus_chars.append(candidate)
                    break

    return "".join(consensus_chars)


def build_hypermut_input(aligned_records: List[SequenceRecord], consensus_name: str, tie_char: str) -> List[SequenceRecord]:
    consensus_seq = consensus_from_alignment(aligned_records, tie_char=tie_char)
    consensus_record = SequenceRecord(header=consensus_name, sequence=consensus_seq)
    return [consensus_record] + aligned_records


# section 6


def auto_preview_start(consensus_seq: str) -> int:
    for idx, base in enumerate(consensus_seq):
        if base != "-":
            return idx
    return 0


def runs_to_graphic_features(seq_window: str) -> List[GraphicFeature]:
    color_map = {
        "A": "#2e7d32",
        "C": "#1565c0",
        "G": "#ef6c00",
        "T": "#6a1b9a",
        "-": "#9e9e9e",
    }
    features: List[GraphicFeature] = []
    i = 0
    while i < len(seq_window):
        char = seq_window[i]
        j = i + 1
        while j < len(seq_window) and seq_window[j] == char:
            j += 1

        run_len = j - i
        label = char if run_len <= 2 else None
        features.append(
            GraphicFeature(
                start=i,
                end=j,
                strand=+1,
                color=color_map.get(char, "#546e7a"),
                label=label,
            )
        )
        i = j
    return features


def write_dna_features_viewer_previews(
    consensus_record: SequenceRecord,
    aligned_records: List[SequenceRecord],
    cfg: PipelineConfig,
) -> Tuple[Path | None, Path | None]:
    if not cfg.write_graphical_previews or plt is None or GraphicRecord is None:
        LOGGER.info("Skipping graphical previews (disabled or dependency unavailable)")
        return None, None

    if cfg.preview_window_start is None:
        start = auto_preview_start(consensus_record.sequence)
    else:
        start = max(0, cfg.preview_window_start)
    size = max(1, cfg.preview_window_size)
    n_show = max(1, cfg.preview_num_sequences)

    consensus_window = consensus_record.sequence[start : start + size]
    consensus_features = runs_to_graphic_features(consensus_window)
    consensus_gr = GraphicRecord(sequence_length=len(consensus_window), features=consensus_features)

    consensus_path = cfg.output_dir / "consensus_preview_dna_features.png"
    fig1, ax1 = plt.subplots(figsize=(max(10, size / 6), 2.8))
    consensus_gr.plot(ax=ax1, with_ruler=True)
    ax1.set_title(f"Consensus Preview ({start}-{start + size - 1})")
    fig1.tight_layout()
    fig1.savefig(consensus_path, dpi=220)
    plt.close(fig1)
    LOGGER.info("Step complete: consensus graphical preview written to %s", consensus_path)

    rows = [consensus_record] + aligned_records[:n_show]
    labels = ["consensus"] + [r.header[:24] for r in aligned_records[:n_show]]

    fig_h = max(4.2, 0.75 * len(rows) + 1.3)
    fig2, axes = plt.subplots(
        nrows=len(rows),
        ncols=1,
        figsize=(max(10, size / 6), fig_h),
        sharex=True,
    )
    if len(rows) == 1:
        axes = [axes]

    for ax, rec, label in zip(axes, rows, labels):
        seq_window = rec.sequence[start : start + size]
        gr = GraphicRecord(sequence_length=len(seq_window), features=runs_to_graphic_features(seq_window))
        gr.plot(ax=ax, with_ruler=False)
        ax.set_ylabel(label, rotation=0, ha="right", va="center")
        ax.set_yticks([])

    axes[-1].set_xlabel(f"Alignment window positions ({start}-{start + size - 1})")
    fig2.suptitle("Consensus + Aligned Sequences (DnaFeaturesViewer)", y=0.995)
    fig2.tight_layout()

    alignment_path = cfg.output_dir / "alignment_preview_dna_features.png"
    fig2.savefig(alignment_path, dpi=220)
    plt.close(fig2)
    LOGGER.info("Step complete: alignment graphical preview written to %s", alignment_path)
    return consensus_path, alignment_path


def base_fill_map() -> dict:
    if PatternFill is None:
        return {}
    return {
        "A": PatternFill(fill_type="solid", start_color="FFB7E1CD", end_color="FFB7E1CD"),
        "C": PatternFill(fill_type="solid", start_color="FFB3D9FF", end_color="FFB3D9FF"),
        "G": PatternFill(fill_type="solid", start_color="FFFFD6A5", end_color="FFFFD6A5"),
        "T": PatternFill(fill_type="solid", start_color="FFE2C7FF", end_color="FFE2C7FF"),
        "-": PatternFill(fill_type="solid", start_color="FFE0E0E0", end_color="FFE0E0E0"),
        "N": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "R": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "Y": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "S": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "W": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "K": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "M": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "B": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "D": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "H": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
        "V": PatternFill(fill_type="solid", start_color="FFFFF3B0", end_color="FFFFF3B0"),
    }


def write_excel_alignment_output(
    consensus_record: SequenceRecord,
    aligned_records: List[SequenceRecord],
    cfg: PipelineConfig,
) -> Path | None:
    if not cfg.write_excel_output or Workbook is None or PatternFill is None:
        LOGGER.info("Skipping Excel output (disabled or openpyxl unavailable)")
        return None

    fills = base_fill_map()
    default_fill = fills.get("N")

    workbook = Workbook()
    ws_consensus = workbook.active
    ws_consensus.title = "Consensus"
    ws_consensus.append(["position", "base"])

    for pos, base in enumerate(consensus_record.sequence, start=1):
        ws_consensus.append([pos, base])
        cell = ws_consensus.cell(row=pos + 1, column=2)
        cell.fill = fills.get(base, default_fill)

    ws_consensus.freeze_panes = "A2"
    ws_consensus.column_dimensions["A"].width = 12
    ws_consensus.column_dimensions["B"].width = 8

    ws_alignment = workbook.create_sheet(title="Alignment")
    header = ["sequence_name"] + list(range(1, len(consensus_record.sequence) + 1))
    ws_alignment.append(header)

    rows = [consensus_record] + aligned_records
    for rec in rows:
        ws_alignment.append([rec.header] + list(rec.sequence))

    ws_alignment.freeze_panes = "B2"
    ws_alignment.column_dimensions["A"].width = 34

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(2, len(consensus_record.sequence) + 2):
            cell = ws_alignment.cell(row=row_idx, column=col_idx)
            base = str(cell.value)
            cell.fill = fills.get(base, default_fill)

    excel_path = cfg.output_dir / cfg.excel_filename
    workbook.save(excel_path)
    LOGGER.info("Step complete: Excel output written to %s", excel_path)
    return excel_path


# section 7

def preview_status_message(cfg: PipelineConfig) -> str:
    if not cfg.write_graphical_previews:
        return "disabled_by_config"
    if plt is None:
        return "matplotlib_not_available"
    if GraphicRecord is None:
        return "dna_features_viewer_not_available"
    return "generated"


def excel_status_message(cfg: PipelineConfig) -> str:
    if not cfg.write_excel_output:
        return "disabled_by_config"
    if Workbook is None or PatternFill is None:
        return "openpyxl_not_available"
    return "generated"

def run_pipeline(cfg: PipelineConfig) -> dict:
    LOGGER.info("Pipeline started")
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    LOGGER.info("Output directory: %s", cfg.output_dir)

    records = read_fasta(cfg.input_fasta)
    if not records:
        raise ValueError(f"No FASTA records found in {cfg.input_fasta}")
    LOGGER.info("Loaded %d input sequences from %s", len(records), cfg.input_fasta)

    total_input_records = len(records)
    if cfg.max_sequences_for_test > 0:
        records = records[: cfg.max_sequences_for_test]
        LOGGER.info(
            "Test mode active: using first %d sequences",
            len(records),
        )
    else:
        LOGGER.info("Test mode disabled: using all input sequences")

    cleaned_records, qc_rows = clean_records(records, cfg)
    if not cleaned_records:
        raise ValueError("All sequences were filtered out during cleaning")
    LOGGER.info("Normalized %d sequences (no filtering policy)", len(cleaned_records))

    qc_path = cfg.output_dir / "qc_cleaning_summary.csv"
    write_qc_csv(qc_rows, qc_path)
    LOGGER.info("Step complete: QC summary written to %s", qc_path)

    cleaned_path = cfg.output_dir / "cleaned_sequences.fasta"
    write_fasta(cleaned_records, cleaned_path)
    LOGGER.info("Step complete: cleaned FASTA written to %s", cleaned_path)

    aligned = is_aligned(cleaned_records)
    aligned_path = cfg.output_dir / "aligned_sequences.fasta"
    if aligned and not cfg.force_mafft_alignment:
        LOGGER.info("Input appears aligned; reusing cleaned sequences as alignment")
        write_fasta(cleaned_records, aligned_path)
        aligned_records = cleaned_records
    else:
        LOGGER.info("Input not aligned (or force flag enabled); starting alignment")
        tmp_input = cfg.output_dir / "_mafft_input.fasta"
        aligned_records = align_with_mafft(cleaned_records, cfg, tmp_input=tmp_input, out_path=aligned_path)
        if tmp_input.exists():
            tmp_input.unlink()
            LOGGER.info("Removed temporary MAFFT input file %s", tmp_input)

    LOGGER.info("Alignment ready with %d sequences", len(aligned_records))

    hypermut_records = build_hypermut_input(
        aligned_records=aligned_records,
        consensus_name=cfg.consensus_name,
        tie_char=cfg.consensus_tie_char,
    )
    hypermut_input_path = cfg.output_dir / "hypermut3_input_with_consensus_first.fasta"
    write_fasta(hypermut_records, hypermut_input_path)
    LOGGER.info("Step complete: Hypermut3 input FASTA written to %s", hypermut_input_path)

    consensus_record = hypermut_records[0]
    consensus_png_path, alignment_png_path = write_dna_features_viewer_previews(
        consensus_record=consensus_record,
        aligned_records=aligned_records,
        cfg=cfg,
    )
    excel_path = write_excel_alignment_output(
        consensus_record=consensus_record,
        aligned_records=aligned_records,
        cfg=cfg,
    )

    run_manifest = {
        "input_fasta": str(cfg.input_fasta),
        "n_input_records": total_input_records,
        "n_input_records_used": len(records),
        "max_sequences_for_test": cfg.max_sequences_for_test,
        "n_cleaned_records": len(cleaned_records),
        "input_already_aligned": aligned,
        "force_mafft_alignment": cfg.force_mafft_alignment,
        "consensus_name": cfg.consensus_name,
        "consensus_tie_char": cfg.consensus_tie_char,
        "qc_csv": str(qc_path),
        "cleaned_fasta": str(cleaned_path),
        "aligned_fasta": str(aligned_path),
        "hypermut_input_fasta": str(hypermut_input_path),
        "preview_status": preview_status_message(cfg),
        "consensus_preview_png": str(consensus_png_path) if consensus_png_path else "not_generated",
        "alignment_preview_png": str(alignment_png_path) if alignment_png_path else "not_generated",
        "excel_status": excel_status_message(cfg),
        "excel_output": str(excel_path) if excel_path else "not_generated",
    }

    manifest_path = cfg.output_dir / "pipeline_manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["key", "value"])
        for k, v in run_manifest.items():
            writer.writerow([k, v])

    LOGGER.info("Step complete: pipeline manifest written to %s", manifest_path)
    LOGGER.info("Pipeline finished successfully")

    return run_manifest


def configure_logging(log_level: str) -> None:
    level_name = (log_level or "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def main() -> None:
    project_root = Path(__file__).resolve().parent
    cfg = PipelineConfig(
        input_fasta=project_root / "data" / "hiv-db-complete-genome-problematic-nucleutides.fasta",
        output_dir=project_root / "data" / "processed",
        max_sequences_for_test=0,  # Set 0 to use all sequences.
        force_mafft_alignment=False,
        mafft_binary="mafft",
        mafft_threads=-1,
        consensus_name="Subtype_A_Consensus",
        consensus_tie_char="ACGT",  # Tie handling now stays within A/C/G/T.
        preview_window_start=None,
        preview_window_size=120,
        preview_num_sequences=12,
        write_graphical_previews=True,
        write_excel_output=True,
        excel_filename="consensus_and_alignment.xlsx",
        log_level="INFO",
    )

    configure_logging(cfg.log_level)

    results = run_pipeline(cfg)
    print("Pipeline completed. Key outputs:")
    for key, value in results.items():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()
